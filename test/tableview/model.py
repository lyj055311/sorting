from PyQt5.QtCore import QStringListModel
from PyQt5.QtGui import QStandardItemModel,QStandardItem
import openpyxl
class Mod:
    def __init__(self):
        self.mod2()
    @property
    def mod1(self):
        self.qlist=['按钮一','按钮二','按钮三','按钮四']
        slm=QStringListModel()
        slm.setStringList(self.qlist)
        return slm
    def mod2(self):
        wb=openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\学生\总表\总表.xlsx')
        ws=wb.active
        self.mod_2=QStandardItemModel(ws.max_row-1,4)
        self.mod_2.setHorizontalHeaderLabels(['id','姓名','性别','学籍号'])

        for i in range(ws.max_row-1):
            a=QStandardItem(str(i+1))
            self.mod_2.setItem(i,0,a)
            self.mod_2.setItem(i,1,QStandardItem(ws.cell(row=i+2,column=3).value))
            self.mod_2.setItem(i,2,QStandardItem(ws.cell(row=i+2,column=4).value))
            self.mod_2.setItem(i,3,QStandardItem(ws.cell(row=i+2,column=5).value))



