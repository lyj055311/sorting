# import openpyxl
# wb = openpyxl.load_workbook('实验1.xlsx')
# sheet_1 = wb.get_active_sheet()
# sheet_2 = wb.get_sheet_by_name('Sheet1')
#在表格中插入空行，空列
    # insert_rows(self,idx, amount=1)和insert_cols(self,idx, amount=1)；
    #参数self，对象是worksheet,这里创建的实例对象为sheet_1
    #参数idx：插入的行/列位置
    #参数amount：插入行/列数量
# sheet_1.insert_rows(3)

# sheet_1.insert_rows(3,30)

# sheet_1.insert_cols(4,10)


#删除行和列
#delete_rows(self, idx, amount=1)和delete_cols(self, idx, amount=1)
# 参数idx：删除的行/列位置
# 参数amount：删除行/列数量
# sheet_1.delete_cols(4,3)
#
# print(openpyxl.utils.column_index_from_string('G'))
# sheet_1.delete_cols(openpyxl.utils.column_index_from_string('G'))

#移动单元格内容move_range(cell_range, rows=0, cols=0）
#参数cell_range:需要移动的单元格范围
#参数rows>0时，向下移动，row<0时，向上移动
#参数cols>0,向右移动，cols<0时，向左移动
# sheet_1.move_range('B5:D8',rows=-30,cols=0)

# sheet_1.move_range("A11:B12",rows=-1,cols=1)  #选中区域移动


#openpyxl提供numpy和pandas交互的接口
#Pandas基于numpy的基础,主要适用数据结构是一维数据与DataFrame(二维数据)，
#这两种数据结构足以处理金融、统计、社会科学、工程等领域里的大多数典型用例。
import  pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import  numpy as np
# dates = pd.date_range('20220101', periods=6)
# df = pd.DataFrame(np.random.randn(6, 4), index=dates, columns=list('ABCD'))
# print(df)   #创建好一个DataFrame数据
# for r in dataframe_to_rows(df, index=True, header=True):
#     sheet_2.append(r)   #写入到excel表格中
# wb.save('实验1.xlsx')
# df = pd.DataFrame(sheet_2.values)  #从表格中读取数据，转换成pandas中的DataFrame数据类型
# print(df)



'''###################图表#################################
openpyxl中包含：
1、Area Charts  面积图          函数方法
     2D Area Charts  2D面积图  AreaChart   
     3D Area Charts  3D面积图  AreaChart3D
2、Bar and Column Charts  条形图/柱状图BarChart
     Vertical, Horizontal and Stacked Bar Charts 水平，垂直和堆积条形图
     3D Bar Charts     3D条形图BarChart3D
3、Bubble Charts         气泡图BubbleChart
4、 Line Charts           折线图
      Line Charts       折线图 LineChart
      3D Line Charts    3D折线图LineChart3D
5、Scatter Charts        散点图 ScatterChart
6、Pie Charts            饼图
      Pie Charts        饼图      PieChart
      Projected Pie Charts  投影饼图ProjectedPieChart
      3D Pie Charts     3D饼图    PieChart3D
7、Doughnut Charts         圆环图DoughnutChart
8、Radar Charts            雷达图RadarChart
9、Stock Charts            股票图StockChart
10、Surface charts         曲面图SurfaceChart/SurfaceChart3D'''

'''面积图'''
# import openpyxl
# from openpyxl.chart import AreaChart,Reference,Series,ScatterChart
#
# wb = openpyxl.load_workbook('实验1.xlsx')
# sheet_2 = wb.get_sheet_by_name('Sheet2')
#
# ac = AreaChart()                #创建面积图对象
# ac.title = "股票价格趋势图"    #表标题
# ac.style = 2                   #图表样式选择
# ac.x_axis.title = '7月份'     #x轴的名称
# ac.y_axis.title = '价格'      #y轴的名称
#                              #多组数据时，可以实现多组数据展示
# data = Reference(sheet_2, min_col=2, min_row=2,max_col=3,max_row=32)
# ac.add_data(data, titles_from_data=True) #from_rows:从数据行中设置数据标头，titles_from_data：数据说明可从数据中获取
# x_labes = Reference(sheet_2, min_col=1, min_row=2, max_row=32)
# ac.set_categories(x_labes)     #设置类别/x轴的值
# sheet_2.add_chart(ac, "E1")
# wb.save('实验1.xlsx')


'''3D面积图'''
# import openpyxl
# from openpyxl.chart import AreaChart3D,Reference
#
# wb = openpyxl.load_workbook('实验1.xlsx')
# sheet_3 = wb.get_sheet_by_name('Sheet3')
#
# ac3 = AreaChart3D()
# ac3.title = "股票3D图"
# ac3.style = 7
# ac3.x_axis.title = '7月份'
# ac3.y_axis.title = '价格'
#
# x_labes = Reference(sheet_3, min_col=1, min_row=2, max_row=7)
# data = Reference(sheet_3, min_col=2, min_row=1, max_col=3, max_row=7) #这里只写到了7
# ac3.add_data(data, titles_from_data=True)
# ac3.set_categories(x_labes)
#
# sheet_3.add_chart(ac3, "E10")
#
# wb.save("实验1.xlsx")

'''条形图 BarChart|限定x轴和y轴的范围'''
# import openpyxl
# from openpyxl.chart import BarChart,Reference
#
# wb = openpyxl.load_workbook('实验1.xlsx')
# sheet_3 = wb.get_sheet_by_name('Sheet3')
# bc = BarChart()           #创建条形图对象
# bc.title = "条形图"         #图标题
# bc.style = 2              #图样式
# bc.x_axis.title = '7月份'   #x轴标题
# bc.y_axis.title = '价格'    #y轴标题
# bc.x_axis.scaling.min=0     #限定x轴的最小范围为0
# bc.x_axis.scaling.max = 52  #限定x轴的最大范围为32
# bc.y_axis.scaling.min = 0   #限定y的最小范围为0
# bc.y_axis.scaling.max = 1   #限定y轴的最大范围为3
# data = Reference(sheet_3,min_col=2,min_row=1,max_col=3,max_row=32)
# bc.add_data(data,titles_from_data=True)
# x_lables = Reference(sheet_3,min_col=1,min_row=2,max_row=32)
# bc.set_categories(x_lables)
# sheet_3.add_chart(bc,'E1')
# wb.save('实验1.xlsx')

#3D条形图
# import openpyxl
# from openpyxl.chart import BarChart3D,Reference
# wb = openpyxl.load_workbook('实验1.xlsx')
# sheet_3 = wb.get_sheet_by_name('Sheet3')
# bc3 = BarChart3D()
# bc3.title = "条形图"
# bc3.style = 2
# bc3.x_axis.title = '7月份'
# bc3.y_axis.title = '价格'
# data = Reference(sheet_3,min_col=2,min_row=1,max_col=3,max_row=5)
# bc3.add_data(data,titles_from_data=True)
# x_lables = Reference(sheet_3,min_col=1,min_row=2,max_row=5)
# bc3.set_categories(x_lables)
# sheet_3.add_chart(bc3,'E1')
# wb.save('实验1.xlsx')



#指定x轴和y轴方向 ||散点图ScatterChart
# import openpyxl
from openpyxl.chart import ScatterChart,Series,Reference

# wb = openpyxl.load_workbook('实验1.xlsx')
# sheet_4 = wb.get_sheet_by_name("Sheet4")
# sc1 = ScatterChart()
# sc1.title = '**新材料在不同温度下的电压'
# sc1.x_axis.title = '温度'
# sc1.y_axis.title = '电压'
# x_data = Reference(sheet_4,min_col=1,min_row=2,max_row=12)
# y_data = Reference(sheet_4,min_col=2,min_row=1,max_row=12)
# s = Series(y_data, xvalues=x_data,title_from_data=True)  #将数据集合在一起
# sc1.append(s)
#
# sc2 =ScatterChart()    #增加第二个图
# sc2.title = '**新材料在不同温度下的电压'
# sc2.x_axis.title = '温度'
# sc2.y_axis.title = '电压'
#
# sc2.append(s)
# sc2.x_axis.scaling.orientation = "maxMin"   #从大到小
# sc2.y_axis.scaling.orientation = "minMax"   #从小到大
# sheet_4.add_chart(sc1,'D1')
# sheet_4.add_chart(sc2,'D17')
# wb.save('实验1.xlsx')

#add_data:AreaChart,AreaChart3D,BarChart,BarChart3D,LineChart,LineChart3D,PieChart,ProjectedPieChart,DoughnutChart,RadarChart,StockChart,SurfaceChart
#Series:BubbleChart,ScatterChart,



'''项目：
    **公司20年的营业收入统计为条形图，每一个月份为统计一个条形图
    例如样例
    1、获取目标文件夹下面的每一个文件名（目的：读取每一个文件）
'''

import openpyxl
from openpyxl.chart import BarChart,Reference
import os
for file_name in os.listdir('./excel项目-##公司15年营收'):
    # print(file_name)
    ex_file_name = './excel项目-##公司15年营收/'+file_name
    # print(ex_file_name)
    ex_file = openpyxl.load_workbook(ex_file_name)
    sheet_names = ex_file.get_sheet_names()         #获取工作簿中的所有工作表名
    for sheet_name in sheet_names:
        sheet_file = ex_file.get_sheet_by_name(sheet_name)   #读取到对应工作表
        bc = BarChart()     #创建条形图
        bc.title = file_name+sheet_name
        bc.x_axis.title = '日期'
        bc.y_axis.title = '营收额'
        data = Reference(sheet_file, min_col=2, min_row=1, max_col=4, max_row=31)
        bc.add_data(data,titles_from_data=True)
        x_lables = Reference(sheet_file,min_col=1,min_row=2,max_row=31)
        bc.set_categories(x_lables)
        sheet_file.add_chart(bc,'E5')
        ex_file.save(ex_file_name)


'''单元格样式设置openpyxl.styles
    字体设置 Font-->设置字体名，字体大小，字体颜色，斜体,下划线...
    单元格填充PatternFill-->填充单元颜色
    单元格的边框Border
    单元格对齐Alignment
'''


'''Border单元格的边框设置
Border(left=Side(), right=Side(), top=Side(),
       bottom=Side(), diagonal=Side(), diagonal_direction=None,
       vertical=None, horizontal=None, diagonalUp=False, diagonalDown=False,
       outline=True, start=None, end=None)
       其中left、right、top、bottom分别对应单元格左框、右框、顶框和下框   

Side:用于边框样式的选项.注意:如果不指定边框样式,则其他属性将没有效果！
Side(style=None, color=None, border_style=None):
       其中style和 border_style相同，值可为：'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'
     
'''
# import openpyxl
# from openpyxl.styles import Border,Side,PatternFill
#
# ex_file = openpyxl.load_workbook('实验1.xlsx')
# sheet_7 = ex_file.get_sheet_by_name('Sheet7')
#
# th = Side(style="thin", color="000000")   #创建一个Side对象，其中框类型为细，颜色RGB为000000
# db = Side(style="double", color="ff0000") #创建一个Side对象，其中框类型为双线，颜色RGB为ff0000
# sheet_7['G3'].border = Border(top=db)     #对'G3'单元格的边框顶部设置为db
# sheet_7['G5'].border = Border(top=db, left=th, right=th, bottom=db)#对'G5'单元格的边框顶部设置为db，左边框为th，有边框为th，底边框为db
# ex_file.save('实验1.xlsx')

'''整个表格实现边框'''
# th = Side(style="thin", color="000000")
# db = Side(style="thick", color="000000")   #thick粗线
# line_num = 1
# for line in   sheet_7['A1':'D31']:
#     for cell in line:
#         if line_num ==1 :
#             cell.fill = PatternFill('solid',fgColor="00CDCD")
#             cell.border = Border(left=th, right=th, top=th,bottom=db)
#         else :
#             cell.border = Border(left=th, right=th, top=th, bottom=th)
#     line_num += 1
# ex_file.save('实验1.xlsx')


'''项目-->在**公司20年的营业表格中，添加表格框和填充颜色，与上面的样式相同'''
# import openpyxl
# from openpyxl.styles import Border,Side,PatternFill
# import os
#
# th = Side(style="thin", color="000000")
# db = Side(style="thick", color="000000")   #thick粗线
# for file_name in os.listdir('./excel项目-##公司15年营收'):
#     ex_file_name = './excel项目-##公司15年营收/'+file_name
#     ex_file = openpyxl.load_workbook(ex_file_name)
#     sheet_names = ex_file.get_sheet_names()         #获取工作簿中的所有工作表名
#     for sheet_name in sheet_names:
#         line_num = 1
#         sheet_file = ex_file.get_sheet_by_name(sheet_name)   #读取到对应工作表
#         for line in sheet_file['A1':'D31']:
#             for cell in line:
#                 if line_num == 1:
#                     cell.fill = PatternFill('solid', fgColor="00CDCD")
#                     cell.border = Border(left=th, right=th, top=th, bottom=db)
#                 else:
#                     cell.border = Border(left=th, right=th, top=th, bottom=th)
#             line_num += 1
#         ex_file.save(ex_file_name)



'''Alignment单元格对齐设置
 Alignment(horizontal=None, vertical=None,
           textRotation=0, wrapText=None, shrinkToFit=None, indent=0, relativeIndent=0,
           justifyLastLine=None, readingOrder=0, text_rotation=None,
           wrap_text=None, shrink_to_fit=None, mergeCell=None)
horizontal的值可以为:"general", "left", "center", "right", "fill", "justify", "centerContinuous","distributed"
vertical的值可以为 "top", "center", "bottom", "justify", "distributed"
'''
# import openpyxl
# from openpyxl.styles import Alignment
# ex_file = openpyxl.load_workbook('实验1.xlsx')
# sheet_8 = ex_file.get_sheet_by_name('Sheet8')
# sheet_8['G3'].alignment = Alignment(horizontal='general')
# sheet_8['G4'].alignment = Alignment(horizontal='left')
# # sheet_8['G5'].alignment = Alignment(horizontal='center')
# # sheet_8['G6'].alignment = Alignment(horizontal='right')
# # sheet_8['G7'].alignment = Alignment(horizontal='fill')
# # sheet_8['G8'].alignment = Alignment(horizontal='justify')
# # sheet_8['G9'].alignment = Alignment(horizontal='centerContinuous')
# # sheet_8['G10'].alignment = Alignment(horizontal='distributed')
# #
# # sheet_8['I3'].alignment = Alignment(horizontal='general')
# # sheet_8['I4'].alignment = Alignment(horizontal='left')
# # sheet_8['I5'].alignment = Alignment(horizontal='center')
# # sheet_8['I6'].alignment = Alignment(horizontal='right')
# # sheet_8['I7'].alignment = Alignment(horizontal='fill')
# # sheet_8['I8'].alignment = Alignment(horizontal='justify')
# # sheet_8['I9'].alignment = Alignment(horizontal='centerContinuous')
# # sheet_8['I10'].alignment = Alignment(horizontal='distributed')
#
# # sheet_8['G13'].alignment = Alignment(vertical = 'top')
# # sheet_8['G14'].alignment = Alignment(vertical = 'center')
# # sheet_8['G15'].alignment = Alignment(vertical = 'bottom')
# # sheet_8['G16'].alignment = Alignment(vertical = 'justify')
# sheet_8['G17'].alignment = Alignment(vertical = 'distributed')
# #
# ex_file.save('实验1.xlsx')




'''排序和筛选：要添加筛选器，请定义一个范围，然后添加排序条件：
add_filter_column(col_id, vals, blank=False) #筛选条件
    参数col_id：对指定列设置筛选条件，
            数据类型为数值，0表示筛选区域中第一列
    参数vals：  筛选内容，
            数据类型为列表，对应列的选择内容
    参数blank： 空白设置，如果为True，将选择空白行，默认不选择

add_sort_condition(ref, descending=False)  #排序
    参数ref：选择需要排序的区域
    descending：选择排序升序或降序，默认升序
'''
# import openpyxl
#
# ex_file = openpyxl.load_workbook('实验1.xlsx')
# sheet_9 = ex_file.get_sheet_by_name('Sheet9')
# sheet_9.auto_filter.ref = 'B1:D9'   #创建筛选器
# sheet_9.auto_filter.add_filter_column(0, ['北京','安徽'],blank=True)  #在筛选中添加筛选元素
# sheet_9.auto_filter.add_sort_condition("D2:D8",descending= True)  #排序
# ex_file.save('实验1.xlsx')



''' #################表格#######################

Table(displayName = None,ref=None,tableStyleInfo = None)
    参数displayName:创建表格名，数据类型为字符串，
        注意：表格名不能重复相同，必须唯一！
    参数ref：选择创建表格区域
    参数tableStyleInfo:创建表格样式
TableStyleInfo(name=None,showFirstColumn=None,showLastColumn=None,
                 showRowStripes=None,showColumnStripes=None)
    参数name：设置表格样式；例如TableStyleMedium8
    参数showFirstColumn：设置第一列颜色填充
    参数showLastColumn:设置最后一列颜色填充
    参数showRowStripes:设置表格中行条纹
    参数showColumnStripes:设置表格中列条纹
'''
# import openpyxl
# from openpyxl.worksheet.table import Table, TableStyleInfo
#
# ex_file = openpyxl.load_workbook('实验1.xlsx')
# sheet_9 = ex_file.get_sheet_by_name('Sheet9')
#
# style = TableStyleInfo(name="TableStyleMedium10", showFirstColumn=True,
#                        showLastColumn=True, showRowStripes=False, showColumnStripes=True)
# tab = Table(displayName = 'Table9',ref="M12:P20",tableStyleInfo = style)
# sheet_9.add_table(tab)
# ex_file.save("实验1.xlsx")


'''保护:   1、保护工作簿    2、保护工作表
保护工作簿WorkbookProtection(workbookPassword=None,
            lockStructure = None,lockWindows= None)
    参数:workbookPassword设置工作簿密码，数据类型为字符串
    参数lockStructure:选择结构选项，数据类型为bool
    参数lockWindows:选择窗口选项，数据类型为bool
    
保护工作表SheetProtection(sheet=False, objects=False, scenarios=False,
                     formatCells=True, formatRows=True, formatColumns=True,
                 insertColumns=True, insertRows=True, insertHyperlinks=True,
                 deleteColumns=True, deleteRows=True, selectLockedCells=False,
                 selectUnlockedCells=False, sort=True, autoFilter=True, pivotTables=True,
                 password=None, algorithmName=None, saltValue=None, spinCount=None, hashValue=None)
    参数password：设置工作表密码，数据类型为字符串
          
'''
# import openpyxl
# from openpyxl.workbook.protection import  WorkbookProtection
# from openpyxl.worksheet.protection import  SheetProtection
# ex_file = openpyxl.load_workbook('实验2.xlsx')
# # ex_file.security = WoteorkbookPrction(workbookPassword='123123',lockStructure = True,lockWindows= True)
# sheet_10 = ex_file.get_sheet_by_name('Sheet10')
# sheet_10.protection = SheetProtection(password = '123456789')
# ex_file.save('实验2.xlsx')








