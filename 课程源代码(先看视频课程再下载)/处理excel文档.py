import openpyxl

'''load_workbook()获取工作簿对象'''
# file_1 = openpyxl.load_workbook('实验1.xlsx')
#注意：次文件需为当前代码所在目录里面

'''get_sheet_names() 获取对应工作簿中所有表名的列表'''
# print(file_1.get_sheet_names())
# first_sheet = file_1.active
'''get_sheet_by_name()从工作簿中取得工作表对象'''
# sheet_1 = file_1.get_sheet_by_name('表单2')  #
# print(sheet_1,type(sheet_1))

'''首次excel表格打开时出现的工作表'''
# first_Sheet = file_1.get_active_sheet()
# print(first_Sheet)

'''获取工作表内容'''
# print(sheet_1['A1'])
# print(sheet_1['A1'].value)
# print(sheet_1['B2'].value)  #大小写均可
# print(sheet_1['c3'].value)
# print(sheet_1['B4'].value)
#当前时间：CTRL+；+shift
'''使用cell方法获取工作表内容'''  #row  :行  colunn：列
# print(sheet_1.cell(row = 1,column = 1))  #列使用数字表示，并不使用ABC...
# print(sheet_1.cell(row = 1,column = 1).value)
# for i in range(1,5):#遍历指定位置的cell内容
#     print(sheet_1.cell(row = i,column = 2).value)

'''max_row和 max_column属性获取表的大小'''
# a= sheet_1.max_row
# b = sheet_1.max_column
# print('行数为：{}；列数为：{}'.format(a,b))


'''打印指定列或行的内容'''
# for i in range(a):  #打印第二列的所有内容
#     print(sheet_1.cell(row = i+1,column = 3).value)
#注：   1、excel中的行数从1开始，列数从A开始
#       2、如果有空的cell，显示None

'''列字母与数字之间的转换'''
# print(openpyxl.utils.get_column_letter(100))
# print(openpyxl.utils.column_index_from_string('CV'))

'''对表格切片'''
# print(sheet_1['B2':'C4']) #返回元组，元组子元组为一行
# 读取切片的内容
# for line in sheet_1['B2':'C4']:  #line是一行
#     print(line)
#     for one_cell in line: #one_cell是一个表格cell
#         print(one_cell.value)

'''获取指定行和列'''
# print(list(sheet_1.columns)[0])
# print(list(sheet_1.rows)[1])
# for one_cell in list(sheet_1.columns)[1]:
#     print(one_cell.value)


'''创建excel文件Workbook()'''
# ex_file = openpyxl.Workbook()
# print(ex_file.get_sheet_names())   #默认有一个工作表
# sheet_1 = ex_file.get_active_sheet()

'''修改工作表名称title属性'''
# sheet_1

# print(ex_file.get_sheet_names())

'''保存文件到本地save()'''
# ex_file.save('excel表格数据3.xlsx')

'''在cell中输入数据，与字典存入数据相似'''
# sheet_1['A2'] = '第一次写入数据'
# print(sheet_1['A2'])


'''任务1：
    在B1:C4区间内填充 '第一次写入' '''

# for line in sheet_1['B1:C4']:
#     for one_cell in line:
#         ]]]]]]]]]]]';# 次写入'
#cell.coordinate属性获取单元格的行列信息
#cell.row属性获取单元格的行信息
#cell.column属性获取单元格的列信息
# ex_file.save('excel表格数据3.xlsx')

'''增加工作表create_sheet();参数title，index
   注：以ex_file为对象'''
# ex_file.create_sheet('')
# print(ex_file.get_sheet_names())
# ex_file.create_sheet('新数据表3',index=1)
# ex_file.save('excel表格数据3.xlsx')  #新的数据覆盖旧数据


'''删除工作表remove_sheet()'''
# ex_file.remove_sheet(ex_file.get_sheet_by_name('新数据表3'))
# ex_file.save('excel表格数据3.xlsx')





'''项目---淘宝医用口罩数据整理入excel文件
任务1：
任务1：在当前目录的项目1文件夹中创建一个excel文件
任务2：修改excel工作表名为：口罩供应商
任务3：文件表头为：商品名；	价格；
任务4：将txt文本的内容按照表头复制到excel文件中


'''
# import openpyxl
# import os
# file_path = os.path.join(os.getcwd(),'项目1','医用口罩表.xlsx')
#
# ex_file = openpyxl.Workbook()  #创建一个文件对象
# sheet_1 = ex_file.get_active_sheet()
# sheet_1.title = '口罩供应商'#工作表名为：口罩供应商
# sheet_1['A1'] = '商品名'
# sheet_1['B1'] = '价格'
# #
# file_path_1 = os.path.join(os.getcwd(),'项目1','医用口罩淘宝.txt')#打开txt文本
# file  = open(file_path_1,'r',encoding='utf-8')
# file_info = file.readlines()
# #
# goods = [] #将需要用到的数据保存在列表goods中
# for i in range(0,len(file_info),4):
#     goods.append(file_info[i+2].strip('\n'))
#     for j in range(len(file_info[i+1])):
#         if file_info[i+1][j] == '.':
#             goods.append(float(file_info[i + 1][1:j+3]))
#             break
# file.close()
# print(goods,len(goods))
# #
# # #将数据写入excel文件中
# end_num = 'B'+str(len(goods)//2 +1) #计算出最后一个cell的行列数
# i = 0
# for line in sheet_1['A2':end_num]:
#     for one_cell in line:
#         sheet_1[one_cell.coordinate] = goods[i]
#         i += 1
# #
# ex_file.save(file_path)  #在当前目录的项目1文件夹中创建一个excel文件


'''项目升级---多个淘宝医用口罩数据整理入excel文件'''
# import openpyxl
# import os
# #
# file_path = os.path.join(os.getcwd(),'项目2','医用口罩表.xlsx')
# ex_file = openpyxl.Workbook()  #创建一个文件对象
# sheet_1 = ex_file.get_active_sheet()
# sheet_1.title = '口罩供应商'#工作表名为：口罩供应商
# sheet_1['A1'] = '商品名'
# sheet_1['B1'] = '价格'
# #
# pro_path = os.path.join(os.getcwd(),'项目2')
# goods = []
# for addr in os.listdir(pro_path):
#     one_file_path = os.path.join(os.getcwd(), '项目2', addr)
#     print(one_file_path)
#     file  = open(one_file_path,'r',encoding='utf-8')
#     file_info = file.readlines()
#     for i in range(0,len(file_info),4):
#         goods.append(file_info[i+2].strip('\n'))
#         for j in range(len(file_info[i+1])):
#             if file_info[i+1][j] == '.':
#                 goods.append(float(file_info[i + 1][1:j+3]))
#                 break
#     file.close()
# print(goods,len(goods))
# #
# end_num = 'B'+str(len(goods)//2+1)
# i = 0
# for line in sheet_1['A2':end_num]:
#     for one_cell in line:
#         sheet_1[one_cell.coordinate] = goods[i]
#         i += 1
# ex_file.save(file_path)  #在当前目录的项目1文件夹中创建一个excel文件

#模板套



# import openpyxl
# from openpyxl.styles import Font,PatternFill
#
# ex_file = openpyxl.load_workbook('实验1.xlsx')
# sheet_1 = ex_file.get_active_sheet()
# # '''Font:字体设置'''
# sheet_1['A4'] = '字体'
# sheet_1['A4'].font = Font(name='幼圆')
# #
# sheet_1['A5'] = '我是有颜色的字体'
# sheet_1['A5'].font = Font(color='8A2BE2') #RGB值
# #
# sheet_1['A6'] = '我是斜体字体'
# sheet_1['A6'].font = Font(italic=True)
# #
# sheet_1['A7'] = '有颜色斜体'
# sheet_1['A7'].font = Font(color= 'DA70D6',italic=True)
# #
# sheet_1['A8'] = '字体大小'
# sheet_1['A8'].font = Font(size=18,underline = 'single',italic=True)
# #
# # '''填充PatternFill'''
# sheet_1['B6'] = '填充'
# sheet_1['B6'].fill = PatternFill('solid',fgColor="FFDDDD")
# ex_file.save('实验1.xlsx')

# '''公式求和'''
# sheet_1['F6'] = '=SUM(F1:F5)'
# ex_file.save('excel表格数据3.xlsx')

'''读取excel文件中的值
    1、读取公式
    2、读取最终的值'''
# ex_file = openpyxl.load_workbook('excel表格数据3.xlsx')
# sheet_1 = ex_file.get_active_sheet()
# print(sheet_1['F6'].value)
#
# ex_file = openpyxl.load_workbook('excel表格数据3.xlsx',data_only=True)
# sheet_1 = ex_file.get_active_sheet()
# print(sheet_1['F6'].value)
#注：必须要手动保存，使得excel计算出公式的值，才能获取结果
#对于复杂的excel公式，可以使用python代码来替代，
# python代码更加直观

'''设置行高列宽'''
#row_dimensions 和 column_dimensions 属性
# sheet_1.row_dimensions[8].height = 170

# sheet_1.column_dimensions['A'].width = 70
#
#行高1个数值表示 1/72 英寸，大约0.35mm
#列宽1个数值表示一个字符大小

'''合并单元格'''
# sheet_1.merge_cells('C9:E11')
# sheet_1['C9'] = '合并'  #写入合并的第一个单元格位置
# ex_file.save('excel表格数据3.xlsx')

'''拆分单元格'''
# sheet_1.unmerge_cells('C9:E11')
# ex_file.save('excel表格数据3.xlsx')



'''冻结窗口
冻结对应单元格上一行和左边一列'''
# ex_file = openpyxl.load_workbook('excel表格数据3.xlsx')
# sheet_1 = ex_file.get_sheet_by_name('Sheet')
# sheet_1.freeze_panes = 'B2'
# ex_file.save('excel表格数据3.xlsx')


'''图表'''
'''柱状图'''
# values = openpyxl.chart.Reference(sheet_1, min_row = 2, min_col = 3, max_row = 2, max_col = 10)
# chart = openpyxl.chart.()#chart.title
# chart.title = '第一次创建表格'
# chart.add_data(values)  #数据添加到柱状图中
# sheet_1.add_chart(chart,'K2')  #将柱状图添加到sheet_1
# ex_file.save('excel表格数据3.xlsx')
#1、读取数据，
# 2、创建图表对象
# 3、将数据添加到图表
# 4、将图标添加到对应的表里面
#文档：https://openpyxl.readthedocs.io/en/stable/index.html